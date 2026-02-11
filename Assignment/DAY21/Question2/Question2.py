#! this is with pandas

# import pandas as pd

# df = pd.read_excel("Assignment\DAY21\Question2\sales_data.xlsx", sheet_name="2025")

# df["Total"] = df["Quantity"] * df["Price"]
# print(df)
# df.to_excel("sales_summary.xlsx",index=False)


#! this is now only with openpyxl not with pandas

from openpyxl import load_workbook

wb = load_workbook("Assignment/DAY21/Question2/sales_data.xlsx")

ws = wb["2025"]

ws["D1"] = "Total"

for row in range(2, ws.max_row + 1):
    quantity = ws[f"B{row}"].value
    price = ws[f"C{row}"].value
    ws[f"D{row}"] = quantity * price

wb.save("sales1_summary.xlsx")
